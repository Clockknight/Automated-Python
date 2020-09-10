import os
import sys
import csv
import openpyxl
from openpyxl import workbook

def main():
    for root, dir, file in os.walk(os.getcwd(), topdown = False):#Check for all files in CWD
        for fileName in file:
            if fileName[-5:] == '.xlsx':#If the file is a .xlsx file, run the convert method on it
                convert(root, fileName)#Pass the whole file name through, and the root path



def convert(root, fileTitle):
    source = os.path.join(root, fileTitle)#The source/xlsx file is here, at root\\fileTitle
    csvBase = fileTitle[:-5]#csvBase is the name of the file with no extensions

    book = openpyxl.load_workbook(source)#Begin declaring excel related variables
    sheets = len(book.sheetnames)#Find the amount of sheets in the workbook

    for a in range(0, sheets):
        sheet = book.worksheets[a]#Get information of the sheet as the program iterates through
        sheetName = book.sheetnames[a]
        sheetWidth = sheet.max_column#Find the dimensions of the sheet
        sheetHeight = sheet.max_row

        csvName = csvBase + sheetName + '.csv'
        output = os.path.join(root, csvName)
        target = open(csvName, 'w+', newline='')
        targetWriter = csv.writer(target)


        for y in range(1, sheetHeight+1):
            rowList = []#Clear the rowList variable before moving onto the next row

            for x in range(1, sheetWidth+1):
                rowList.append(sheet.cell(row=y, column=x).value)#Gather values from the sheet's row


            targetWriter.writerow(rowList)

        print('Writing to ' + output + '...')
        target.close()



main()
