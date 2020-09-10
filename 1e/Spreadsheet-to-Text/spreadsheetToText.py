import os
import sys
import openpyxl
from openpyxl import workbook

if len(sys.argv) != 2 or sys.argv[1][-5:] != '.xlsx':
    print('The arguments ran through the command line are invalid! Try running something like\n\tspreadsheetToText.py document.xlsx')

#Take 1 argv: spreadsheet name
    #save the name as sourceName
sourceName = str(sys.argv[1])


book = openpyxl.load_workbook(sourceName)
sheet = book.active

#Take the spreadsheet's max columna and rows
width = sheet.max_column + 1
height = sheet.max_row + 1

#For each row in the sheet, write {cell.value}+'\n' to the text doc.
for i in range(1, width):
    content = '' # Clear content at the start of the loop
    fileName = sourceName[:-5] + ' page ' + str(i) + '.txt'
    file = open(fileName, 'w+')

    for j in range(1, height):
        if sheet.cell(row=j, column=i).value != None:
            content += str(sheet.cell(row=j, column=i).value)
        content += '\n'
    file.write(content)
    print('Saved as ' + fileName + '.')

print('Done.')
