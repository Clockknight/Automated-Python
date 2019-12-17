import os
import sys
import openpyxl
from openpyxl import workbook

#Check rows and columns on the spreadsheet, only actually run if either are 2 or more
    #Take spreadsheet based on argv from cwd
if (len(sys.argv) != 2) or (str(sys.argv[1])[-5:] != '.xlsx'):
    print('This program requires an additional argument to target an .xlsx file.\nTry again with something like:\n\tcellInverter.py file_to_convert.xlsx')
    sys.exit()

    #Load book
book = openpyxl.load_workbook(str(sys.argv[1]))
sheet = book.active

    #Take rows and columns
sheetWidth = sheet.max_column
sheetHeight = sheet.max_row

#Check which dimension is larger, width or  height.
    #Compare, change size to width, then height if height > width
if sheetHeight > sheetWidth:
    sheetWidth = sheetHeight
size = sheetWidth

#For each row, invert cells equal to current row - 1 (row # 5 will invert 4 cells with column 5)
for i in range(1, size):    #For loop, i in dimension
    for j in range(1, i):    #Nested For loop, j in dimension - 1
        #Store values
        sourceValue = sheet.cell(row=i, column=j).value #Take value of cell in (i, j), store as sourceValue
        targetValue = sheet.cell(row=j, column=i).value#Take value of cell in (j, i), store as targetValue
        #Apply stored values
        sheet.cell(row=j, column=i).value = sourceValue
        sheet.cell(row=i, column=j).value = targetValue

#Save
book.save('inverted'+str(sys.argv[1]))
