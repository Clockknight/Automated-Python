import os
import sys
import openpyxl
from openpyxl import workbook


#Take number N from command line

tableSize = 5

#Print out NxN multiplication Tabulates

    #Create book to write data onto
book = Workbook()
sheet = book.active

    #Begin looping through cells, filling out table.
for x in tableSize:
    #Begin writing out values for "label" cells
    sheet.cell(row=1, column=x+1).value = x
    sheet.cell(row=x+1, column=1).value = x
    for y in tableSize:
        sheet.cell(row=x+1, column=y+1).value = x*y #Fill out value of cell at coords (x+1, y+1)

#Bold Row 1 & Column A
boldCell = Font(bold = True)
for cellObject in sheet.columns[0]:
    cellObject.font = boldCell
for cellObject in sheet.rows[0]:
    cellObject.font = boldCell

#Save all data created
book.save('multiplicationTable%dx%d' % (tableSize, tableSize))
