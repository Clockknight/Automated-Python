import os
import sys
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font


#Take number N from command line
if len(sys.argv) != 2:
    print("Sorry, this program requires you to pass an additional argument specifying the dimensions of the multiplication square.\nTry:\n\tmultiplicationTable.py 5 for a 5x5 table, etc...")
    sys.exit()

tableSize = int(sys.argv[1])
print("Creating a %dx%d table..." % (tableSize, tableSize))

#Print out NxN multiplication Tabulates

    #Create book to write data onto
book = openpyxl.Workbook()
book.save('multiplicationTable%dx%d.xlsx' % (tableSize, tableSize))
sheet = book.active

    #Begin looping through cells, filling out table.
for x in range(1, tableSize):
    #Begin writing out values for "label" cells
    sheet.cell(row=1, column=x+1).value = x
    sheet.cell(row=x+1, column=1).value = x
    bold = Font(bold=True)
    sheet.cell(row=1, column=x+1).font = bold
    sheet.cell(row=x+1, column=1).font = bold
    for y in range(1, tableSize):
        sheet.cell(row=x+1, column=y+1).value = x*y #Fill out value of cell at coords (x+1, y+1)

#Save all data created
book.save('multiplicationTable%dx%d.xlsx' % (tableSize, tableSize))
